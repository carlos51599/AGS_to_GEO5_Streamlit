import pandas as pd
import os
import csv
import re
from openpyxl import load_workbook

# ---- PARAMETERS ----
AGS_FILE = "input_file.ags"  # Set your AGS file path
TEMPLATE_FILE = "FieldTestImportTemplate.xlsx"  # Set your template path
OUTPUT_FILE = "Geo5_Import.xlsx"


# ---- AGS PARSING LOGIC (adapted from data_loader.py) ----
def parse_group(content, group_name):
    lines = content.splitlines()
    parsed = list(csv.reader(lines, delimiter=",", quotechar='"'))
    headings = []
    data = []
    in_group = False
    for row in parsed:
        if row and row[0] == "GROUP" and len(row) > 1 and row[1] == group_name:
            in_group = True
            continue
        if in_group and row and row[0] == "HEADING":
            headings = row[1:]
            continue
        if in_group and row and row[0] == "DATA":
            data.append(row[1 : len(headings) + 1])
            continue
        if (
            in_group
            and row
            and row[0] == "GROUP"
            and (len(row) < 2 or row[1] != group_name)
        ):
            break
    df = pd.DataFrame(data, columns=headings)
    return df


# ---- LOAD AGS FILE ----
def load_ags_tables(ags_path):
    with open(ags_path, encoding="utf-8") as f:
        content = f.read()
    geol = parse_group(content, "GEOL")
    point = parse_group(content, "POINT")
    loca = parse_group(content, "LOCA")
    return {"GEOL": geol, "POINT": point, "LOCA": loca}


# ---- DATA LOADING ----
nags_tables = load_ags_tables(AGS_FILE)
df_geol = nags_tables["GEOL"]
df_point = nags_tables["POINT"]
df_loca = nags_tables["LOCA"]

# Convert numeric columns
for col in ["GEOL_DEPTH", "GEOL_BASE"]:
    if col in df_geol.columns:
        df_geol[col] = pd.to_numeric(df_geol[col], errors="coerce")

for col in ["LOCA_NATE", "LOCA_NATN", "LOCA_GL"]:
    if col in df_loca.columns:
        df_loca[col] = pd.to_numeric(df_loca[col], errors="coerce")

# Merge POINT with LOCA for coordinates
if "POINT_ID" in df_point.columns and "LOCA_ID" in df_loca.columns:
    df_point = df_point.merge(
        df_loca,
        left_on="POINT_ID",
        right_on="LOCA_ID",
        how="left",
        suffixes=("", "_loca"),
    )

# ---- LAYER NAME EXTRACTION ----
regex = re.compile(r"^([A-Z]+(?:\s+[A-Z]+)*)")


def extract_layer_name(desc):
    match = regex.match(str(desc))
    return match.group(1).strip() if match and len(match.group(1)) > 1 else ""


if "GEOL_GEOL" in df_geol.columns:
    df_geol["GEOL_GEOL"] = df_geol.apply(
        lambda row: (
            row["GEOL_GEOL"]
            if pd.notna(row.get("GEOL_GEOL", "")) and row["GEOL_GEOL"]
            else extract_layer_name(row.get("GEOL_DESC", ""))
        ),
        axis=1,
    )
else:
    df_geol["GEOL_GEOL"] = df_geol.apply(
        lambda row: extract_layer_name(row.get("GEOL_DESC", "")), axis=1
    )


# ---- SOIL CLASSIFICATION AUTO ASSIGNMENT ----
def auto_classify(geol_leg):
    soil_map = {
        "CLAY": "Clay, fine grained",
        "SAND": "Sand, coarse grained",
        "SILT": "Silt",
        "GRAVEL": "Gravel",
        # Add more mappings as needed
    }
    return soil_map.get(str(geol_leg).upper(), "")


df_geol["Soil Classification"] = df_geol.get("GEOL_LEG", pd.Series("")).apply(
    auto_classify
)

# ---- WRITE TO TEMPLATE ----
wb = load_workbook(TEMPLATE_FILE)
ws_fieldtests = wb["FieldTests"]
ws_layers = wb["Layers"]

# ---- CLEAR AND PREPARE FIELDTESTS SHEET ----
ws_fieldtests.delete_rows(2, ws_fieldtests.max_row)  # keep header

row_idx = 2
for i, row in df_loca.iterrows():
    # Only write rows where LOCA_ID is not empty
    if not row.get("LOCA_ID", ""):
        continue
    ws_fieldtests.cell(row=row_idx, column=1, value=row.get("LOCA_ID", ""))
    ws_fieldtests.cell(row=row_idx, column=2, value="(local set) : Borehole")
    ws_fieldtests.cell(row=row_idx, column=3, value=row.get("LOCA_NATN", ""))
    ws_fieldtests.cell(row=row_idx, column=4, value=row.get("LOCA_NATE", ""))
    ws_fieldtests.cell(row=row_idx, column=5, value="input")
    ws_fieldtests.cell(row=row_idx, column=6, value=row.get("LOCA_GL", ""))
    row_idx += 1

# ---- CLEAR AND PREPARE LAYERS SHEET ----
ws_layers.delete_rows(2, ws_layers.max_row)  # keep header


# --- Assign Soil pattern|Color using adapted VBA logic ---
def to_hex2(val):
    return format(int(val), "02X")


def assign_colors(count_groups):
    # Pastel pink: RGB(255, 209, 220) -> BGR: (220, 209, 255)
    # Pastel yellow: RGB(255, 229, 180) -> BGR: (180, 229, 255)
    colors = []
    if count_groups == 1:
        colors.append("$808080")
        return colors
    for i in range(count_groups):
        if i == 0:
            color_hex = "$808080"  # Top layer: grey
        elif i == count_groups - 1:
            color_hex = "$B4E5FF"  # Bottom layer: pastel yellow (BGR)
        else:
            nIntermediate = count_groups - 2
            j = i - 1
            if nIntermediate > 1:
                f = j / (nIntermediate)
            else:
                f = 0
            # Interpolate between pastel pink and pastel yellow
            R = 255
            G = round(209 + f * (229 - 209))
            B = round(220 + f * (180 - 220))
            color_hex = "$" + to_hex2(B) + to_hex2(G) + to_hex2(R)
        colors.append(color_hex)
    return colors


# Collect and write rows to Layers, assigning colors per borehole
layer_row = 2
if "LOCA_ID" in df_geol.columns and "GEOL_LEG" in df_geol.columns:
    grouped = df_geol.groupby("LOCA_ID")
    for borehole_id, group in grouped:
        # Prepare layer data for this borehole
        borehole_layers = []
        for i, row in group.iterrows():
            geol_top = row.get("GEOL_TOP", None)
            geol_base = row.get("GEOL_BASE", None)
            try:
                thickness = float(geol_base) - float(geol_top)
            except (TypeError, ValueError):
                thickness = ""
            desc = row.get("GEOL_DESC", "")
            borehole_layers.append(
                {
                    "borehole_id": borehole_id,
                    "thickness": thickness,
                    "soil_name": row.get("GEOL_LEG", ""),
                    "desc": desc,
                }
            )
        # Assign colors for this borehole's layers
        colors = assign_colors(len(borehole_layers)) if borehole_layers else []
        # Write rows for this borehole
        for idx, row in enumerate(borehole_layers):
            ws_layers.cell(
                row=layer_row, column=1, value=row["borehole_id"]
            )  # Test name
            ws_layers.cell(row=layer_row, column=2, value=row["thickness"])  # Thickness
            ws_layers.cell(row=layer_row, column=3, value=row["soil_name"])  # Soil name
            ws_layers.cell(
                row=layer_row, column=4, value="GEO_CLAY"
            )  # Soil pattern|Pattern
            ws_layers.cell(
                row=layer_row, column=5, value=colors[idx] if idx < len(colors) else ""
            )  # Soil pattern|Color
            ws_layers.cell(
                row=layer_row, column=6, value="clDefault"
            )  # Soil pattern|Background
            ws_layers.cell(
                row=layer_row, column=7, value="50"
            )  # Soil pattern|Saturation
            ws_layers.cell(
                row=layer_row, column=8, value=row["desc"]
            )  # Layer description
            ws_layers.cell(
                row=layer_row, column=9, value=""
            )  # EN ISO 14688-1 Classification (blank)
            layer_row += 1

# ---- SAVE OUTPUT ----
wb.save(OUTPUT_FILE)
print(f"AGS parsed and exported to {OUTPUT_FILE}")
