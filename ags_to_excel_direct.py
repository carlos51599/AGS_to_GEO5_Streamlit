import pandas as pd
import csv
import re

from openpyxl import load_workbook

# ---- SET YOUR FILE PATHS HERE ----
AGS_FILE = r"C:\Users\dea29431.RSKGAD\OneDrive - Rsk Group Limited\Documents\Geotech\AGS to GEO5 Import\AGS_to_GEO5_Streamlit\FLRG - 2025-05-20 1711 - Preliminary data - 4.ags"  # <-- Set your AGS file path
TEMPLATE_FILE = r"C:\Users\dea29431.RSKGAD\OneDrive - Rsk Group Limited\Documents\Geotech\AGS to GEO5 Import\AGS_to_GEO5_Streamlit\FieldTestImportTemplate.xlsx"  # <-- Set your template path
OUTPUT_FILE = r"C:\Users\dea29431.RSKGAD\OneDrive - Rsk Group Limited\Documents\Geotech\AGS to GEO5 Import\Geo5_ImportDirect.xlsx"  # <-- Set your output Excel path


def parse_group(content, group_name):
    lines = content.splitlines()
    parsed = list(csv.reader(lines, delimiter=",", quotechar='"'))
    headings = []
    data = []
    in_group = False
    for row in parsed:
        if row and row[0] == "GROUP" and len(row) > 1 and row[1] == group_name:
            in_group = True
            continue
        if in_group and row and row[0] == "HEADING":
            headings = row[1:]
            continue
        if in_group and row and row[0] == "DATA":
            data.append(row[1 : len(headings) + 1])
            continue
        if (
            in_group
            and row
            and row[0] == "GROUP"
            and (len(row) < 2 or row[1] != group_name)
        ):
            break
    df = pd.DataFrame(data, columns=headings)
    return df


def assign_colors(count_groups):
    # Pastel pink: RGB(255, 209, 220) -> BGR: (220, 209, 255)
    # Pastel yellow: RGB(255, 229, 180) -> BGR: (180, 229, 255)
    def to_hex2(val):
        return format(int(val), "02X")

    colors = []
    if count_groups == 1:
        colors.append("$808080")
        return colors
    for i in range(count_groups):
        if i == 0:
            color_hex = "$808080"  # Top layer: grey
        elif i == count_groups - 1:
            color_hex = "$B4E5FF"  # Bottom layer: pastel yellow (BGR)
        else:
            nIntermediate = count_groups - 2
            j = i - 1
            if nIntermediate > 1:
                f = j / (nIntermediate)
            else:
                f = 0
            R = 255
            G = round(209 + f * (229 - 209))
            B = round(220 + f * (180 - 220))
            color_hex = "$" + to_hex2(B) + to_hex2(G) + to_hex2(R)
        colors.append(color_hex)
    return colors


def main():
    with open(AGS_FILE, encoding="utf-8") as f:
        content = f.read()
    geol = parse_group(content, "GEOL")
    point = parse_group(content, "POINT")
    loca = parse_group(content, "LOCA")
    # Convert numeric columns (force to float, coerce errors to NaN)
    for col in ["GEOL_TOP", "GEOL_BASE", "GEOL_DEPTH"]:
        if col in geol.columns:
            geol[col] = pd.to_numeric(geol[col], errors="coerce")
    for col in ["LOCA_NATE", "LOCA_NATN", "LOCA_GL"]:
        if col in loca.columns:
            loca[col] = pd.to_numeric(loca[col], errors="coerce")
    # Write to Excel template
    wb = load_workbook(TEMPLATE_FILE)
    ws_fieldtests = wb["FieldTests"]
    ws_layers = wb["Layers"]
    # Clear FieldTests
    ws_fieldtests.delete_rows(2, ws_fieldtests.max_row)
    row_idx = 2
    for i, row in loca.iterrows():
        if not row.get("LOCA_ID", ""):
            continue
        ws_fieldtests.cell(row=row_idx, column=1, value=row.get("LOCA_ID", ""))
        ws_fieldtests.cell(row=row_idx, column=2, value="(local set) : Borehole")
        ws_fieldtests.cell(row=row_idx, column=3, value=row.get("LOCA_NATN", ""))
        ws_fieldtests.cell(row=row_idx, column=4, value=row.get("LOCA_NATE", ""))
        ws_fieldtests.cell(row=row_idx, column=5, value="input")
        ws_fieldtests.cell(row=row_idx, column=6, value=row.get("LOCA_GL", ""))
        row_idx += 1
    # Clear Layers
    ws_layers.delete_rows(2, ws_layers.max_row)
    layer_row = 2
    if "LOCA_ID" in geol.columns and "GEOL_LEG" in geol.columns:
        grouped = geol.groupby("LOCA_ID")
        for borehole_id, group in grouped:
            borehole_layers = []
            for i, row in group.iterrows():
                geol_top = row.get("GEOL_TOP", None)
                geol_base = row.get("GEOL_BASE", None)
                try:
                    thickness = float(geol_base) - float(geol_top)
                except (TypeError, ValueError):
                    thickness = ""
                desc = row.get("GEOL_DESC", "")
                borehole_layers.append(
                    {
                        "borehole_id": borehole_id,
                        "thickness": thickness,
                        "soil_name": row.get("GEOL_LEG", ""),
                        "desc": desc,
                    }
                )
            colors = assign_colors(len(borehole_layers)) if borehole_layers else []
            for idx, row in enumerate(borehole_layers):
                ws_layers.cell(row=layer_row, column=1, value=row["borehole_id"])
                ws_layers.cell(row=layer_row, column=2, value=row["thickness"])
                ws_layers.cell(row=layer_row, column=3, value=row["soil_name"])
                ws_layers.cell(row=layer_row, column=4, value="GEO_CLAY")
                # Soil pattern|Color: assign a default color (grey) for now
                ws_layers.cell(row=layer_row, column=5, value="$808080")
                ws_layers.cell(row=layer_row, column=6, value="clDefault")
                # Soil pattern|Saturation: ensure this is a number, not text
                ws_layers.cell(row=layer_row, column=7, value=50)
                ws_layers.cell(row=layer_row, column=8, value=row["desc"])
                ws_layers.cell(row=layer_row, column=9, value="")
                layer_row += 1
    wb.save(OUTPUT_FILE)
    print(f"Exported AGS to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
