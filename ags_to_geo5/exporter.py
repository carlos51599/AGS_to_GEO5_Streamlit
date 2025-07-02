import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def extract_layer_name(desc):
    regex = re.compile(r"^([A-Z]+(?:\s+[A-Z]+)*)")
    match = regex.match(str(desc))
    return match.group(1).strip() if match and len(match.group(1)) > 1 else ""


def auto_classify(geol_leg):
    soil_map = {
        "CLAY": "Clay, fine grained",
        "SAND": "Sand, coarse grained",
        "SILT": "Silt",
        "GRAVEL": "Gravel",
        # Add more mappings as needed
    }
    return soil_map.get(str(geol_leg).upper(), "")


def export_to_excel(df_geol, df_loca, df_abbr, template_path, output_path):
    # Prepare data
    # Ensure numeric columns for all possible top/base naming conventions
    for col in ["GEOL_TOP", "GEOL_BASE", "GEOL_DEPTH"]:
        if col in df_geol.columns:
            df_geol[col] = pd.to_numeric(df_geol[col], errors="coerce")
    for col in ["LOCA_NATE", "LOCA_NATN", "LOCA_GL"]:
        if col in df_loca.columns:
            df_loca[col] = pd.to_numeric(df_loca[col], errors="coerce")
    # Write to Excel
    wb = load_workbook(template_path)
    ws_fieldtests = wb["FieldTests"]
    ws_layers = wb["Layers"]
    ws_fieldtests.delete_rows(2, ws_fieldtests.max_row)
    row_idx = 2
    for i, row in df_loca.iterrows():
        # Only write rows with a LOCA_ID and at least one coordinate
        if (
            not row.get("LOCA_ID", "")
            and pd.isna(row.get("LOCA_NATN", None))
            and pd.isna(row.get("LOCA_NATE", None))
            and pd.isna(row.get("LOCA_GL", None))
        ):
            continue
        ws_fieldtests.cell(row=row_idx, column=1, value=row.get("LOCA_ID", ""))
        ws_fieldtests.cell(row=row_idx, column=2, value="EN - Standard: Borehole")
        ws_fieldtests.cell(row=row_idx, column=3, value=row.get("LOCA_NATN", ""))
        ws_fieldtests.cell(row=row_idx, column=4, value=row.get("LOCA_NATE", ""))
        ws_fieldtests.cell(row=row_idx, column=5, value="input")
        ws_fieldtests.cell(row=row_idx, column=6, value=row.get("LOCA_GL", ""))
        row_idx += 1
    ws_layers.delete_rows(2, ws_layers.max_row)
    layer_row = 2

    # Group by LOCA_ID, then assign color logic per borehole
    def to_hex2(val):
        return format(int(val), "02X")

    def assign_colors(count_groups):
        # Pastel pink: RGB(255, 209, 220) -> BGR: (220, 209, 255)
        # Pastel yellow: RGB(255, 229, 180) -> BGR: (180, 229, 255)
        colors = []
        if count_groups == 1:
            colors.append("$808080")
            return colors
        for i in range(count_groups):
            if i == 0:
                color_hex = "$808080"  # Top layer: grey
            elif i == count_groups - 1:
                color_hex = "$B4E5FF"  # Bottom layer: pastel yellow (BGR)
            else:
                nIntermediate = count_groups - 2
                j = i - 1
                if nIntermediate > 1:
                    f = j / (nIntermediate)
                else:
                    f = 0
                R = 255
                G = round(209 + f * (229 - 209))
                B = round(220 + f * (180 - 220))
                color_hex = "$" + to_hex2(B) + to_hex2(G) + to_hex2(R)
            colors.append(color_hex)
        return colors

    if "LOCA_ID" in df_geol.columns and "GEOL_LEG" in df_geol.columns:
        for borehole_id, bh_data in df_geol.groupby("LOCA_ID"):
            # Each row in bh_data is a layer
            layers = []
            for leg, layer_data in bh_data.groupby("GEOL_LEG"):
                start_depth = (
                    layer_data["GEOL_TOP"].min()
                    if "GEOL_TOP" in layer_data.columns
                    else None
                )
                end_depth = (
                    layer_data["GEOL_BASE"].max()
                    if "GEOL_BASE" in layer_data.columns
                    else None
                )
                thickness = (
                    end_depth - start_depth
                    if start_depth is not None and end_depth is not None
                    else None
                )
                desc = "; ".join(
                    layer_data.get("GEOL_DESC", pd.Series("")).astype(str).tolist()
                )
                # Soil name from ABBR_DESC
                soil_name = leg
                if (
                    df_abbr is not None
                    and "ABBR_CODE" in df_abbr.columns
                    and "ABBR_DESC" in df_abbr.columns
                ):
                    abbr_match = df_abbr[df_abbr["ABBR_CODE"] == str(leg)]
                    if not abbr_match.empty:
                        soil_name = abbr_match["ABBR_DESC"].iloc[0]
                layers.append(
                    {
                        "borehole_id": borehole_id,
                        "thickness": thickness,
                        "soil_name": soil_name,
                        "desc": desc,
                    }
                )
            colors = assign_colors(len(layers)) if layers else []
            for idx, row in enumerate(layers):
                ws_layers.cell(row=layer_row, column=1, value=row["borehole_id"])
                ws_layers.cell(row=layer_row, column=2, value=row["thickness"])
                ws_layers.cell(row=layer_row, column=3, value=row["soil_name"])
                ws_layers.cell(row=layer_row, column=4, value="GEO_CLAY")
                ws_layers.cell(
                    row=layer_row,
                    column=5,
                    value=colors[idx] if idx < len(colors) else "",
                )
                ws_layers.cell(row=layer_row, column=6, value="clDefault")
                ws_layers.cell(row=layer_row, column=7, value=50)
                ws_layers.cell(row=layer_row, column=8, value=row["desc"])
                ws_layers.cell(row=layer_row, column=9, value="")
                layer_row += 1
    wb.save(output_path)
